import openpyxl
import pprint
wb = openpyxl.load_workbook('D:\\taive\\maTran.xlsx')
print(wb.sheetnames)
# ['Sheet1', 'Sheet2']
sheet = wb['Sheet1']
cellA2 = sheet['A2']
cellB3 = sheet['B3']
cell = sheet['A2']
print(cell.value)
# one


cell = sheet.cell(row=2, column=1)
print(cell.value)
# one
